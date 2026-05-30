"""
影院数据自动清洗脚本 - 核心处理逻辑。
可独立运行（python auto_clean.py），也可被 bot 导入调用 process_data()。
"""
import os
import re
import glob
import numpy as np
import pandas as pd
import openpyxl
from datetime import datetime, timezone, timedelta

CST = timezone(timedelta(hours=8))

def beijing_now():
    return datetime.now(CST)


# ============================================================
# Column matching
# ============================================================
def find_column(df, *keyword_groups):
    """Find column name matching any group of keywords.
    Each group is a list of keywords. Tries ALL-match first, then ANY-match."""
    for keywords in keyword_groups:
        if not keywords:
            continue
        for col in df.columns:
            if all(kw in str(col).strip() for kw in keywords):
                return col
        for col in df.columns:
            if any(kw in str(col).strip() for kw in keywords):
                return col
    return None


# ============================================================
# File identification
# ============================================================
def identify_files(work_dir):
    xlsx_files = glob.glob(os.path.join(work_dir, "*.xlsx")) + glob.glob(os.path.join(work_dir, "*.xls"))
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith("~$")]

    file1 = file2 = file3 = None
    for f in xlsx_files:
        name_no_ext = os.path.splitext(os.path.basename(f))[0]
        if name_no_ext.endswith("-落"):
            file1 = f
        elif name_no_ext.startswith("影城明细-"):
            file2 = f

    for f in xlsx_files:
        if f != file1 and f != file2:
            file3 = f
            break

    if not file1: raise FileNotFoundError("未找到文件1（<电影名称>-落.xlsx）")
    if not file2: raise FileNotFoundError("未找到文件2（影城明细-<电影名称>.xlsx）")
    if not file3: raise FileNotFoundError("未找到文件3")

    movie_name = os.path.splitext(os.path.basename(file1))[0].removesuffix("-落")
    return file1, file2, file3, movie_name


# ============================================================
# Process File 3
# ============================================================
def process_file3(file3_path, target_paipian_value, total_cost, output_dir):
    df3 = pd.read_excel(file3_path)

    target_col = find_column(df3, ["目标排片"])
    status_col = find_column(df3, ["筛选状态", "状态"])
    price_col = find_column(df3, ["单价", "元"], ["单价"])
    if not target_col: raise ValueError(f"未找到目标排片列。可用: {list(df3.columns)}")
    if not status_col: raise ValueError(f"未找到状态列。可用: {list(df3.columns)}")
    if not price_col: raise ValueError(f"未找到单价列。可用: {list(df3.columns)}")

    df3[target_col] = target_paipian_value
    df3 = df3[~df3[status_col].isin(["已撤回", "已驳回"])]

    nv = pd.to_numeric(df3[price_col], errors='coerce')
    price_sum = nv.sum()
    price_count = nv.count()

    actual_consumption = (price_sum / total_cost * 100) if total_cost != 0 else 0

    f3_basename = os.path.splitext(os.path.basename(file3_path))[0]
    f3_output = os.path.join(output_dir, f"{f3_basename}_已处理.xlsx")
    df3.to_excel(f3_output, index=False)

    return df3, actual_consumption, price_count, f3_output


# ============================================================
# Process File 1 - Sheet1
# ============================================================
def process_file1_sheet1(file1_path, file2_path, df3, movie_name):
    df1 = pd.read_excel(file1_path, sheet_name="Sheet1")
    df2 = pd.read_excel(file2_path)

    id_col_1 = find_column(df1, ["影城ID"], ["ID"])
    if not id_col_1: raise ValueError(f"文件1未找到影城ID列: {list(df1.columns)}")
    id_col_2 = find_column(df2, ["影城ID"], ["ID"])
    if not id_col_2: raise ValueError(f"文件2未找到影城ID列: {list(df2.columns)}")

    # File 2 lookup columns
    movie_field_col = find_column(df2, [movie_name, "场次数"], [movie_name, "场"], [movie_name])
    total_field_col = find_column(df2, ["影城总场次数"], ["总场次数"], ["场次数"])

    # File 1 target columns
    col_zhuawawa = find_column(df1, ["7.27", "抓娃娃", "场次数"], ["抓娃娃", "场次数"], ["抓娃娃"])
    col_yinyuan = find_column(df1, ["7.27", "影院", "场次数"], ["影院", "场次数"])
    col_hezuo_bili = find_column(df1, ["合作比例"], ["比例"])
    col_shifou_hezuo = find_column(df1, ["是否合作"], ["合作"])

    # File 3 lookup columns
    target_col_3 = find_column(df3, ["目标排片"])
    cinema_id_3 = find_column(df3, ["影城ID"], ["影院ID"], ["ID"])

    # Merge from File 2
    lookup_cols_2 = [id_col_2]
    lookup_rename_2 = {}
    if movie_field_col:
        lookup_cols_2.append(movie_field_col)
        lookup_rename_2[movie_field_col] = "_zhuawawa_lookup"
    if total_field_col:
        lookup_cols_2.append(total_field_col)
        lookup_rename_2[total_field_col] = "_yinyuan_lookup"

    lookup_2 = df2[lookup_cols_2].copy().drop_duplicates(subset=id_col_2)
    lookup_2.rename(columns=lookup_rename_2, inplace=True)
    df1 = df1.merge(lookup_2, left_on=id_col_1, right_on=id_col_2, how="left")
    if id_col_2 in df1.columns and id_col_2 != id_col_1:
        df1.drop(columns=[id_col_2], inplace=True)

    if col_zhuawawa and "_zhuawawa_lookup" in df1.columns:
        df1[col_zhuawawa] = df1["_zhuawawa_lookup"]
        df1.drop(columns=["_zhuawawa_lookup"], inplace=True)
    if col_yinyuan and "_yinyuan_lookup" in df1.columns:
        df1[col_yinyuan] = df1["_yinyuan_lookup"]
        df1.drop(columns=["_yinyuan_lookup"], inplace=True)

    # Merge from File 3
    if col_hezuo_bili and target_col_3 and cinema_id_3:
        lookup_3 = df3[[cinema_id_3, target_col_3]].copy().drop_duplicates(subset=cinema_id_3)
        lookup_3.rename(columns={target_col_3: "_hezuo_bili_lookup"}, inplace=True)
        df1 = df1.merge(lookup_3, left_on=id_col_1, right_on=cinema_id_3, how="left")
        if cinema_id_3 in df1.columns and cinema_id_3 != id_col_1:
            df1.drop(columns=[cinema_id_3], inplace=True)
        if "_hezuo_bili_lookup" in df1.columns:
            df1[col_hezuo_bili] = df1["_hezuo_bili_lookup"]
            df1.drop(columns=["_hezuo_bili_lookup"], inplace=True)

    # Fill NAs
    for col in [col_zhuawawa, col_yinyuan, col_hezuo_bili]:
        if col and col in df1.columns:
            df1[col] = pd.to_numeric(df1[col], errors='coerce').fillna(0)

    # 是否合作
    if col_shifou_hezuo and col_hezuo_bili and col_hezuo_bili in df1.columns:
        df1[col_shifou_hezuo] = df1[col_hezuo_bili].apply(
            lambda x: "否" if pd.to_numeric(x, errors='coerce') == 0 else "是")

    # Recompute formula columns (pandas reads them as 0 because they're Excel formulas)
    col_ting = find_column(df1, ["厅数"], ["厅"])
    col_weikai = col_hezuo_yingkai = col_yingcheng_yingkai = col_remaining = None
    if col_ting and col_ting in df1.columns:
        ting_numeric = pd.to_numeric(df1[col_ting], errors='coerce').fillna(0)
        bili_numeric = pd.to_numeric(df1[col_hezuo_bili], errors='coerce').fillna(0) if col_hezuo_bili in df1.columns else 0
        zhuawawa_numeric = pd.to_numeric(df1[col_zhuawawa], errors='coerce').fillna(0)
        yinyuan_numeric = pd.to_numeric(df1[col_yinyuan], errors='coerce').fillna(0)

        # 影城应开场次数 = ROUNDUP(厅数 * 5.5)
        col_yingcheng_yingkai = find_column(df1, ["影城应开场次数"]) or "影城应开场次数"
        df1[col_yingcheng_yingkai] = np.ceil(ting_numeric * 5.5)

        # 合作应开场次数 = ROUNDUP(厅数 * 5.5 * 合作比例)
        col_hezuo_yingkai = find_column(df1, ["合作应开场次数"]) or "合作应开场次数"
        df1[col_hezuo_yingkai] = np.ceil(ting_numeric * 5.5 * bili_numeric)

        # 影院未开场次数 = 影城应开场次数 - 7.27影院场次数
        col_weikai = find_column(df1, ["影院未开场次数"]) or "影院未开场次数"
        df1[col_weikai] = df1[col_yingcheng_yingkai] - yinyuan_numeric

        # 7.27抓娃娃剩余 = 合作应开场次数 - 7.27抓娃娃场次数
        col_remaining = find_column(df1, ["7.27", "抓娃娃", "剩余"], ["抓娃娃", "剩余"]) or "7.27抓娃娃剩余"
        df1[col_remaining] = df1[col_hezuo_yingkai] - zhuawawa_numeric

    col_faxing = find_column(df1, ["发行归属"])

    return df1, {
        "id_col_1": id_col_1,
        "col_7_27_zhuawawa": col_zhuawawa,
        "col_7_27_yinyuan": col_yinyuan,
        "col_hezuo_bili": col_hezuo_bili,
        "col_shifou_hezuo": col_shifou_hezuo,
        "col_weikai": col_weikai,
        "col_hezuo_yingkai": col_hezuo_yingkai,
        "col_yingcheng_yingkai": col_yingcheng_yingkai,
        "col_remaining": col_remaining,
        "col_faxing": col_faxing,
    }


# ============================================================
# 落位 helpers
# ============================================================
def read_luowei_from_original(file1_path):
    wb = openpyxl.load_workbook(file1_path, data_only=True)
    ws = wb["落位"]
    result = {
        "b6": ws["B6"].value,
        "f9": ws["F9"].value,
        "f16": ws["F16"].value,
    }
    wb.close()
    return result


def compute_luowei_percentages(df1, col_info, d8_pct):
    shifou_col = col_info["col_shifou_hezuo"]
    zhuawawa_col = col_info["col_7_27_zhuawawa"]
    yinyuan_col = col_info["col_7_27_yinyuan"]

    col_weikai = col_info.get("col_weikai")
    col_hezuo_yingkai = col_info.get("col_hezuo_yingkai")
    col_yingcheng_yingkai = col_info.get("col_yingcheng_yingkai")
    col_faxing = col_info.get("col_faxing")

    pct = d8_pct / 100.0
    mask_yes = df1[shifou_col] == "是"
    mask_no = (df1[shifou_col] == "否")
    if col_faxing and col_faxing in df1.columns:
        mask_no = mask_no & (df1[col_faxing] == "单体")

    def safe_sum(col, mask):
        if col and col in df1.columns:
            return pd.to_numeric(df1.loc[mask, col], errors='coerce').fillna(0).sum()
        return 0.0

    # F9 (分日新增比例测算)
    d7 = safe_sum(col_hezuo_yingkai, mask_yes)
    e7 = safe_sum(col_yingcheng_yingkai, mask_yes)
    d8 = safe_sum(col_weikai, mask_no) * pct + safe_sum(zhuawawa_col, mask_no)
    e8 = safe_sum(col_yingcheng_yingkai, mask_no)
    f9 = (d7 + d8) / (e7 + e8) if (e7 + e8) != 0 else 0.0

    # F16 (实际开场数量测算)
    d14 = safe_sum(col_hezuo_yingkai, mask_yes)
    e14 = safe_sum(col_yingcheng_yingkai, mask_yes)
    zhuawawa_no = safe_sum(zhuawawa_col, mask_no)
    yinyuan_no = safe_sum(yinyuan_col, mask_no)
    e15 = safe_sum(col_yingcheng_yingkai, mask_no)
    d15 = (zhuawawa_no / yinyuan_no * e15) if yinyuan_no != 0 else 0.0
    f16 = (d14 + d15) / (e14 + e15) if (e14 + e15) != 0 else 0.0

    return f9, f16


def save_file1_and_write_d8(file1_path, df1_sheet1, output_dir, d8_pct):
    import xlwings as xw

    app = xw.App(visible=False)
    try:
        wb = app.books.open(file1_path)

        if "Sheet1" in [s.name for s in wb.sheets]:
            ws = wb.sheets["Sheet1"]
            ws.used_range.clear_contents()
            ws.range("A1").value = [str(c) for c in df1_sheet1.columns]
            data = df1_sheet1.where(df1_sheet1.notna(), None).values.tolist()
            if data:
                ws.range("A2").value = data

        if "落位" in [s.name for s in wb.sheets]:
            wb.sheets["落位"].range("D8").formula = (
                f'=GETPIVOTDATA("求和项:影院未开场次数",$K$3,"是否合作","否")*{d8_pct/100}'
                f'+GETPIVOTDATA("求和项:7.27抓娃娃场次数",$K$3,"是否合作","否")'
            )

        out = os.path.join(output_dir, os.path.basename(file1_path))
        wb.save(out)
        wb.close()
        return out
    finally:
        app.quit()


# ============================================================
# Helpers
# ============================================================
def parse_date_from_b6(b6_val):
    if b6_val is None: return str(beijing_now().day)
    m = re.search(r'(\d+)日', str(b6_val))
    if m: return m.group(1)
    m = re.search(r'(\d+)', str(b6_val))
    if m: return m.group(1)
    return str(beijing_now().day)


# ============================================================
# Generate 文案
# ============================================================
def generate_wenan1(movie_name, backend_consume, actual_consumption, prev_actual, price_count, b6_val):
    month = beijing_now().month
    day = parse_date_from_b6(b6_val)
    growth = actual_consumption - prev_actual
    return (f"{month}月{day}日《{movie_name}》单体任务后台消耗{backend_consume}%，"
            f"实际消耗{actual_consumption:.1f}%（含分区未审批）,"
            f"较上一时段增长{growth:+.1f}%,"
            f"实际提报{price_count}家。")


def generate_wenan2(movie_name, df1, col_info, luowei_data):
    now = beijing_now()
    date_str = f"{now.month}月{now.day}日"
    time_str = "11点" if now.hour < 12 else "17点"
    day_num = parse_date_from_b6(luowei_data["b6"])
    prefix = f"{now.month}月{day_num}日《{movie_name}》"

    shifou_col = col_info["col_shifou_hezuo"]
    zhuawawa_col = col_info["col_7_27_zhuawawa"]
    yinyuan_col = col_info["col_7_27_yinyuan"]

    hezuo_yes = df1[df1[shifou_col] == "是"] if shifou_col in df1.columns else df1
    total_coop = len(hezuo_yes)

    z_numeric = pd.to_numeric(hezuo_yes[zhuawawa_col], errors='coerce').fillna(0) if zhuawawa_col in hezuo_yes.columns else pd.Series([0])
    y_numeric = pd.to_numeric(hezuo_yes[yinyuan_col], errors='coerce').fillna(0) if yinyuan_col in hezuo_yes.columns else pd.Series([0])

    opened = int((z_numeric != 0).sum())
    zero_open = int((z_numeric == 0).sum())
    already_open = int(z_numeric.sum())

    col_hezuo_yingkai = col_info.get("col_hezuo_yingkai")
    should_open = int(pd.to_numeric(hezuo_yes[col_hezuo_yingkai], errors='coerce').fillna(0).sum()) if col_hezuo_yingkai in hezuo_yes.columns else 0

    col_remaining = col_info.get("col_remaining")
    not_open_count = 0
    if col_remaining and col_remaining in hezuo_yes.columns:
        r_numeric = pd.to_numeric(hezuo_yes[col_remaining], errors='coerce').fillna(0)
        not_open_count = int(r_numeric[r_numeric > 0].sum())

    opened_not_started = int(((y_numeric > 0) & (z_numeric == 0)).sum())

    return (
        f"截止{date_str}{time_str}，{prefix}提报单体影城开场情况：\n"
        f"单体目前合作影城数：{total_coop}家\n"
        f"合作已开{movie_name}影城数：{opened}家\n"
        f"合作0开{movie_name}影城数：{zero_open}家\n"
        f"合作影城应开{movie_name}场次数：{should_open}场\n"
        f"合作影城已开{movie_name}场次数：{already_open}场\n"
        f"合作影城未开{movie_name}场次数：{not_open_count}场\n"
        f"合作影城已开场未开{movie_name}影城数：{opened_not_started}家\n"
        f"\n"
        f"辛苦老师们针对未开《{movie_name}》影城做重点开场沟通，利用平台活动做抓手加速开场[加油]"
    )


def generate_wenan3(movie_name, b6_val, luowei_data):
    day_num = parse_date_from_b6(b6_val)
    month = beijing_now().month
    date_formatted = f"{month}.{day_num}"

    def parse_num(val):
        if val is None: return 0.0
        try: return float(str(val).strip().replace("%", ""))
        except: return 0.0

    def to_pct_str(val):
        num = parse_num(val)
        return f"{num*100:.1f}%"

    f9_num = parse_num(luowei_data["f9"])
    f16_num = parse_num(luowei_data["f16"])

    if f9_num <= f16_num:
        low_str, high_str = to_pct_str(f9_num), to_pct_str(f16_num)
        low_label = "分日新增比例测算"
        high_label = "实际开场数量测算"
    else:
        low_str, high_str = to_pct_str(f16_num), to_pct_str(f9_num)
        low_label = "实际开场数量测算"
        high_label = "分日新增比例测算"

    return (
        f"《{movie_name}》 {date_formatted}\n"
        f"单体落位预估：{low_str}-{high_str}\n"
        f"注：低值按照{low_label}，高值按照{high_label}"
    )


# ============================================================
# Main entry point
# ============================================================
def process_data(work_dir, output_dir, target_paipian, total_cost, backend_consume, prev_actual, d8_pct):
    os.makedirs(output_dir, exist_ok=True)

    file1, file2, file3, movie_name = identify_files(work_dir)
    luowei_data = read_luowei_from_original(file1)

    df3, actual_consumption, price_count, file3_output = process_file3(file3, target_paipian, total_cost, output_dir)
    df1_sheet1, col_info = process_file1_sheet1(file1, file2, df3, movie_name)

    file1_output = save_file1_and_write_d8(file1, df1_sheet1, output_dir, d8_pct)

    f9_val, f16_val = compute_luowei_percentages(df1_sheet1, col_info, d8_pct)
    luowei_data["f9"] = f9_val
    luowei_data["f16"] = f16_val

    return {
        "wenan1": generate_wenan1(movie_name, backend_consume, actual_consumption, prev_actual, price_count, luowei_data["b6"]),
        "wenan2": generate_wenan2(movie_name, df1_sheet1, col_info, luowei_data),
        "wenan3": generate_wenan3(movie_name, luowei_data["b6"], luowei_data),
        "movie_name": movie_name,
        "file1_output": file1_output,
        "file3_output": file3_output,
        "actual_consumption": actual_consumption,
        "price_count": price_count,
    }
